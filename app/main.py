from datetime import date, datetime, timedelta
from io import BytesIO
import logging
from pathlib import Path
import re
import shutil
import threading
import time
from urllib.parse import urlencode
from collections import defaultdict, deque
from urllib.parse import urlparse

from fastapi import Depends, FastAPI, File, Form, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session, joinedload
from starlette.middleware.sessions import SessionMiddleware

from app.auth import authenticate_user, get_current_user, hash_password, verify_password
from app.db import Base, DATABASE_PATH, PROJECT_ROOT, SessionLocal, engine, get_db
from app.models import AdminActionLog, AppVariable, AuditLog, Location, Provider, RoleEnum, User, Visit
from app.seed import seed_initial_data
from app.services import (
    build_audit_export,
    build_logs_export,
    build_user_import_guide,
    DELAY_NOTE_FIELDS,
    FIELD_LABELS,
    TIME_FIELDS,
    USER_IMPORT_REQUIRED_HEADERS,
    ValidationError,
    build_export,
    current_status,
    day_range,
    delay_note_entries,
    format_dt,
    get_next_field,
    lab_duration_minutes,
    minutes_between,
    other_begin_options,
    other_can_begin,
    other_duration_minutes,
    other_pending_slots,
    now_local,
    override_timestamp,
    set_timestamp,
)

app = FastAPI(title="Patient Cycle Time")
app.add_middleware(
    SessionMiddleware,
    secret_key="replace-this-in-production",
    https_only=False,
    same_site="lax",
    max_age=60 * 60 * 12,
)
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")
MIN_PASSWORD_LENGTH = 8
MAX_PASSWORD_LENGTH = 32
DEFAULT_DAILY_CHECKOUT_GOAL = 6
MAX_MRN_LENGTH = 32
MAX_USERNAME_LENGTH = 64
MAX_SEARCH_LENGTH = 200
MAX_LOCATION_NAME_LENGTH = 120
MAX_PROVIDER_NAME_LENGTH = 120
MAX_VISIT_TYPE_LENGTH = 120
MAX_EXPORT_RANGE_DAYS = 366
MAX_USER_IMPORT_FILE_BYTES = 2 * 1024 * 1024
MAX_USER_IMPORT_ROWS = 2000
SESSION_IDLE_TIMEOUT_SECONDS = 60 * 60
SESSION_ABSOLUTE_TIMEOUT_SECONDS = 60 * 60 * 12
RATE_LIMIT_RULES = {
    ("POST", "/login"): (20, 60),
    ("POST", "/admin/override"): (30, 60),
    ("POST", "/admin/purge-zero-mrn"): (5, 60),
    ("POST", "/admin/backup-db"): (4, 60),
    ("GET", "/admin/logs-export"): (20, 60),
    ("GET", "/admin/audit-export"): (20, 60),
}
_rate_limit_buckets: dict[tuple[str, str, str], deque[float]] = defaultdict(deque)
_rate_limit_lock = threading.Lock()
logger = logging.getLogger("ncf_ptwait")
if not logging.getLogger().handlers:
    logging.basicConfig(level=logging.INFO)


def role_label(role: RoleEnum) -> str:
    return {
        RoleEnum.ADMIN: "Admin",
        RoleEnum.FD: "Front Desk",
        RoleEnum.NURSE: "Nurse",
        RoleEnum.AUDITOR: "Auditor",
    }[role]


templates.env.globals["role_label"] = role_label


def _request_client_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def _passes_rate_limit(request: Request) -> bool:
    key = (request.method.upper(), request.url.path)
    rule = RATE_LIMIT_RULES.get(key)
    if not rule:
        return True
    max_requests, window_seconds = rule
    now = time.time()
    bucket_key = (request.method.upper(), request.url.path, _request_client_ip(request))
    with _rate_limit_lock:
        bucket = _rate_limit_buckets[bucket_key]
        while bucket and now - bucket[0] > window_seconds:
            bucket.popleft()
        if len(bucket) >= max_requests:
            return False
        bucket.append(now)
    return True


def _is_same_origin(request: Request) -> bool:
    origin = request.headers.get("origin") or request.headers.get("referer")
    if not origin:
        return True
    parsed = urlparse(origin)
    host = request.headers.get("host", "")
    return parsed.netloc == host


@app.middleware("http")
async def security_and_observability_middleware(request: Request, call_next):
    start = time.time()
    if request.method.upper() in {"POST", "PUT", "PATCH", "DELETE"} and not _is_same_origin(request):
        return Response("Invalid origin.", status_code=403)
    if not _passes_rate_limit(request):
        return Response("Too many requests. Please try again shortly.", status_code=429)

    response = await call_next(request)
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; img-src 'self' data:; script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline'; object-src 'none'; base-uri 'self'; frame-ancestors 'none'"
    )
    if request.method.upper() != "GET" or response.status_code >= 400:
        duration_ms = int((time.time() - start) * 1000)
        logger.info("%s %s -> %s (%sms)", request.method.upper(), request.url.path, response.status_code, duration_ms)
    return response


def summarize_selected_names(names: list[str], fallback_label: str) -> str:
    if not names:
        return fallback_label
    if len(names) <= 2:
        return ", ".join(names)
    return f"{names[0]} + {len(names) - 1} more"


def format_dt_local_input(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%Y-%m-%dT%H:%M")


def format_dt_minutes(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%Y-%m-%d %H:%M")


def clamp_text(value: str | None, max_length: int) -> str:
    return (value or "").strip()[:max_length]


@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception):
    logger.exception("Unhandled error on %s %s", request.method, request.url.path)
    return Response("Unexpected server error.", status_code=500)


@app.on_event("startup")
def startup() -> None:
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        ensure_user_preference_columns(db)
        ensure_location_columns(db)
        ensure_provider_columns(db)
        ensure_visit_columns(db)
        ensure_app_variables(db)
        ensure_password_change_backfill(db)
        seed_initial_data(db)
    finally:
        db.close()


def set_flash(request: Request, level: str, message: str) -> None:
    request.session["flash"] = {"level": level, "message": message}


def pop_flash(request: Request):
    return request.session.pop("flash", None)


def require_user(request: Request, db: Session) -> User | RedirectResponse:
    now_ts = int(time.time())
    logged_in_at = int(request.session.get("logged_in_at", 0) or 0)
    last_seen_at = int(request.session.get("last_seen_at", 0) or 0)
    if logged_in_at and now_ts - logged_in_at > SESSION_ABSOLUTE_TIMEOUT_SECONDS:
        request.session.clear()
        return RedirectResponse(url="/", status_code=303)
    if last_seen_at and now_ts - last_seen_at > SESSION_IDLE_TIMEOUT_SECONDS:
        request.session.clear()
        return RedirectResponse(url="/", status_code=303)
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    request.session["last_seen_at"] = now_ts
    if user.must_change_password and request.url.path not in {"/force-password", "/logout"}:
        return RedirectResponse(url="/force-password", status_code=303)
    return user


def require_admin(request: Request, db: Session) -> User | RedirectResponse:
    user = require_user(request, db)
    if isinstance(user, RedirectResponse):
        return user
    if user.role != RoleEnum.ADMIN:
        return RedirectResponse(url="/dashboard", status_code=303)
    return user


def log_admin_action(db: Session, acting_user: User, action_name: str, details: str) -> None:
    db.add(
        AdminActionLog(
            action_name=action_name,
            details=details,
            performed_by_user_id=acting_user.id,
        )
    )


def ensure_user_preference_columns(db: Session) -> None:
    columns = {
        row[1]
        for row in db.execute(text("PRAGMA table_info(users)")).fetchall()
    }
    if "preferred_location_id" not in columns:
        db.execute(text("ALTER TABLE users ADD COLUMN preferred_location_id INTEGER"))
    if "preferred_provider_id" not in columns:
        db.execute(text("ALTER TABLE users ADD COLUMN preferred_provider_id INTEGER"))
    if "preferred_location_ids" not in columns:
        db.execute(text("ALTER TABLE users ADD COLUMN preferred_location_ids TEXT"))
    if "preferred_provider_ids" not in columns:
        db.execute(text("ALTER TABLE users ADD COLUMN preferred_provider_ids TEXT"))
    if "is_hidden" not in columns:
        db.execute(text("ALTER TABLE users ADD COLUMN is_hidden INTEGER NOT NULL DEFAULT 0"))
    if "must_change_password" not in columns:
        db.execute(text("ALTER TABLE users ADD COLUMN must_change_password INTEGER NOT NULL DEFAULT 0"))
    if "disable_fancy_effects" not in columns:
        db.execute(text("ALTER TABLE users ADD COLUMN disable_fancy_effects INTEGER NOT NULL DEFAULT 0"))
    if "coins" not in columns:
        db.execute(text("ALTER TABLE users ADD COLUMN coins INTEGER NOT NULL DEFAULT 0"))
    if "daily_checkout_count" not in columns:
        db.execute(text("ALTER TABLE users ADD COLUMN daily_checkout_count INTEGER NOT NULL DEFAULT 0"))
    if "daily_checkout_date" not in columns:
        db.execute(text("ALTER TABLE users ADD COLUMN daily_checkout_date TEXT"))
    db.commit()


def ensure_location_columns(db: Session) -> None:
    columns = {
        row[1]
        for row in db.execute(text("PRAGMA table_info(locations)")).fetchall()
    }
    if "is_hidden" not in columns:
        db.execute(text("ALTER TABLE locations ADD COLUMN is_hidden INTEGER NOT NULL DEFAULT 0"))
    db.commit()


def ensure_provider_columns(db: Session) -> None:
    columns = {
        row[1]
        for row in db.execute(text("PRAGMA table_info(providers)")).fetchall()
    }
    if "is_hidden" not in columns:
        db.execute(text("ALTER TABLE providers ADD COLUMN is_hidden INTEGER NOT NULL DEFAULT 0"))
    db.commit()


def ensure_visit_columns(db: Session) -> None:
    columns = {
        row[1]
        for row in db.execute(text("PRAGMA table_info(visits)")).fetchall()
    }
    if "visit_type" not in columns:
        db.execute(text("ALTER TABLE visits ADD COLUMN visit_type TEXT"))
    if "other_begin_at" not in columns:
        db.execute(text("ALTER TABLE visits ADD COLUMN other_begin_at DATETIME"))
    if "other_end_at" not in columns:
        db.execute(text("ALTER TABLE visits ADD COLUMN other_end_at DATETIME"))
    if "intake_begin_at" not in columns:
        db.execute(text("ALTER TABLE visits ADD COLUMN intake_begin_at DATETIME"))
    if "declined_participation" not in columns:
        db.execute(text("ALTER TABLE visits ADD COLUMN declined_participation INTEGER NOT NULL DEFAULT 0"))
    if "no_show" not in columns:
        db.execute(text("ALTER TABLE visits ADD COLUMN no_show INTEGER NOT NULL DEFAULT 0"))
    if "other_begin_log" not in columns:
        db.execute(text("ALTER TABLE visits ADD COLUMN other_begin_log TEXT"))
    if "other_end_log" not in columns:
        db.execute(text("ALTER TABLE visits ADD COLUMN other_end_log TEXT"))
    if "other_timestamps_json" not in columns:
        db.execute(text("ALTER TABLE visits ADD COLUMN other_timestamps_json TEXT"))
    for column_name in DELAY_NOTE_FIELDS.values():
        if column_name not in columns:
            db.execute(text(f"ALTER TABLE visits ADD COLUMN {column_name} TEXT"))
    db.commit()


def ensure_app_variables(db: Session) -> None:
    existing = db.query(AppVariable).filter(AppVariable.key == "daily_checkout_goal").first()
    if not existing:
        db.add(AppVariable(key="daily_checkout_goal", value=str(DEFAULT_DAILY_CHECKOUT_GOAL)))
        db.commit()


def get_daily_checkout_goal(db: Session) -> int:
    existing = db.query(AppVariable).filter(AppVariable.key == "daily_checkout_goal").first()
    if not existing:
        return DEFAULT_DAILY_CHECKOUT_GOAL
    try:
        return max(0, int((existing.value or "").strip()))
    except ValueError:
        return DEFAULT_DAILY_CHECKOUT_GOAL


def set_daily_checkout_goal(db: Session, goal_value: int) -> None:
    existing = db.query(AppVariable).filter(AppVariable.key == "daily_checkout_goal").first()
    if not existing:
        db.add(AppVariable(key="daily_checkout_goal", value=str(goal_value)))
    else:
        existing.value = str(goal_value)
        db.add(existing)
    db.commit()


def ensure_password_change_backfill(db: Session) -> None:
    marker_key = "password_force_backfill_done"
    marker = db.query(AppVariable).filter(AppVariable.key == marker_key).first()
    if marker:
        return
    users = db.query(User).all()
    for user in users:
        user.must_change_password = True
        db.add(user)
    db.add(AppVariable(key=marker_key, value="1"))
    db.commit()


def password_requirement_errors(password: str) -> list[str]:
    errors: list[str] = []
    if len(password) < MIN_PASSWORD_LENGTH:
        errors.append(f"Must be at least {MIN_PASSWORD_LENGTH} characters.")
    if not re.search(r"[A-Z]", password):
        errors.append("Must include an uppercase letter.")
    if not re.search(r"[a-z]", password):
        errors.append("Must include a lowercase letter.")
    if not re.search(r"[0-9]", password):
        errors.append("Must include a number.")
    if not re.search(r"[^A-Za-z0-9]", password):
        errors.append("Must include a special character.")
    if len(password) > MAX_PASSWORD_LENGTH:
        errors.append(f"Must be {MAX_PASSWORD_LENGTH} characters or fewer.")
    return errors


def sync_user_daily_checkout_state(user: User, db: Session) -> int:
    today_label = date.today().isoformat()
    if user.daily_checkout_date == today_label:
        return user.daily_checkout_count or 0
    user.daily_checkout_date = today_label
    user.daily_checkout_count = 0
    db.add(user)
    db.commit()
    db.refresh(user)
    return 0


def parse_id_csv(value: str | None) -> list[int]:
    if not value:
        return []
    parsed: list[int] = []
    for part in value.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            parsed.append(int(part))
        except ValueError:
            continue
    return parsed


def serialize_id_list(values: list[int]) -> str | None:
    return ",".join(str(value) for value in values) if values else None


def normalize_selected_ids(raw_values: list[int] | None, valid_ids: set[int]) -> list[int]:
    normalized: list[int] = []
    for value in raw_values or []:
        if value in valid_ids and value not in normalized:
            normalized.append(value)
    return normalized


def persisted_selected_ids(csv_value: str | None, fallback_value: int | None, valid_ids: set[int]) -> list[int]:
    parsed = normalize_selected_ids(parse_id_csv(csv_value), valid_ids)
    if parsed:
        return parsed
    if fallback_value in valid_ids:
        return [fallback_value]
    return []


def persist_user_context(user: User, location_ids: list[int], provider_ids: list[int], db: Session) -> None:
    changed = False
    serialized_location_ids = serialize_id_list(location_ids)
    serialized_provider_ids = serialize_id_list(provider_ids)
    first_location_id = location_ids[0] if location_ids else None
    first_provider_id = provider_ids[0] if provider_ids else None

    if user.preferred_location_ids != serialized_location_ids:
        user.preferred_location_ids = serialized_location_ids
        changed = True
    if user.preferred_provider_ids != serialized_provider_ids:
        user.preferred_provider_ids = serialized_provider_ids
        changed = True
    if user.preferred_location_id != first_location_id:
        user.preferred_location_id = first_location_id
        changed = True
    if user.preferred_provider_id != first_provider_id:
        user.preferred_provider_id = first_provider_id
        changed = True
    if changed:
        db.add(user)
        db.commit()
        db.refresh(user)


def build_filter_query(
    location_ids: list[int],
    provider_ids: list[int],
    visit_date: str,
    search: str | None,
    hide_complete: bool,
    location_filter_applied: bool = True,
    provider_filter_applied: bool = True,
) -> str:
    search_value = clamp_text(search, MAX_SEARCH_LENGTH)
    return urlencode(
        [
            *[("location_id", value) for value in location_ids],
            *[("provider_id", value) for value in provider_ids],
            ("visit_date", visit_date),
            ("search", search_value),
            ("hide_complete", str(hide_complete).lower()),
            ("location_filter_applied", str(location_filter_applied).lower()),
            ("provider_filter_applied", str(provider_filter_applied).lower()),
        ]
    )


def dashboard_redirect_url(
    location_ids: list[int],
    provider_ids: list[int],
    visit_date: str,
    search: str | None,
    hide_complete: bool,
    location_filter_applied: bool = True,
    provider_filter_applied: bool = True,
) -> str:
    return "/dashboard?" + build_filter_query(
        location_ids,
        provider_ids,
        visit_date,
        search,
        hide_complete,
        location_filter_applied,
        provider_filter_applied,
    )


def parameters_page_context(request: Request, user: User, db: Session) -> dict:
    return {
        "request": request,
        "current_user": user,
        "users": db.query(User).order_by(User.username.asc()).all(),
        "locations": db.query(Location).order_by(Location.name.asc()).all(),
        "providers": db.query(Provider).order_by(Provider.name.asc()).all(),
        "daily_checkout_goal": get_daily_checkout_goal(db),
        "flash": pop_flash(request),
    }


def visit_matches_dashboard_filters(visit: Visit, user: User, search_text_lower: str, hide_complete: bool) -> bool:
    if hide_complete:
        if visit.no_show or visit.declined_participation:
            return False
        if user.role == RoleEnum.FD and visit.ready_for_clinical_at is not None:
            return False
        if user.role in {RoleEnum.NURSE, RoleEnum.AUDITOR, RoleEnum.ADMIN} and visit.checkout_at is not None:
            return False

    if search_text_lower:
        searchable_text = " ".join(
            [
                visit.mrn or "",
                visit.location.name or "",
                visit.provider.name or "",
                " ".join(note for _, note in delay_note_entries(visit)),
            ]
        ).lower()
        if search_text_lower not in searchable_text:
            return False

    return True


@app.get("/", response_class=HTMLResponse)
def login_page(request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user(request, db)
    if current_user:
        return RedirectResponse(url="/dashboard", status_code=303)
    return templates.TemplateResponse("login.html", {"request": request, "error": None})


@app.post("/login")
def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    user = authenticate_user(db, username=username, password=password)
    if not user:
        return templates.TemplateResponse(
            "login.html", {"request": request, "error": "Invalid username or password."}, status_code=401
        )

    now_ts = int(time.time())
    request.session["user_id"] = user.id
    request.session["logged_in_at"] = now_ts
    request.session["last_seen_at"] = now_ts
    return RedirectResponse(url="/force-password" if user.must_change_password else "/dashboard", status_code=303)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/", status_code=303)


@app.get("/healthz")
def healthcheck():
    return {"status": "ok", "time": datetime.now().isoformat()}


@app.post("/admin/backup-db")
def admin_backup_database(
    request: Request,
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user
    backup_dir = PROJECT_ROOT / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"clinic_cycle_time_{timestamp}.db"
    shutil.copy2(DATABASE_PATH, backup_path)
    log_admin_action(db, user, "database_backup", f"Created backup: {backup_path.name}")
    db.commit()
    set_flash(request, "success", f"Database backup created: {backup_path.name}")
    return RedirectResponse(url="/admin", status_code=303)


@app.get("/force-password", response_class=HTMLResponse)
def force_password_page(request: Request, db: Session = Depends(get_db)):
    user = require_user(request, db)
    if isinstance(user, RedirectResponse):
        return user
    if not user.must_change_password:
        return RedirectResponse(url="/dashboard", status_code=303)
    return templates.TemplateResponse(
        "force_password.html",
        {
            "request": request,
            "current_user": None,
            "flash": pop_flash(request),
            "username": user.username,
            "min_password_length": MIN_PASSWORD_LENGTH,
        },
    )


@app.post("/force-password")
def force_password_update(
    request: Request,
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    db: Session = Depends(get_db),
):
    user = require_user(request, db)
    if isinstance(user, RedirectResponse):
        return user
    if not user.must_change_password:
        return RedirectResponse(url="/dashboard", status_code=303)
    if new_password != confirm_password:
        set_flash(request, "error", "New password and confirmation do not match.")
        return RedirectResponse(url="/force-password", status_code=303)
    requirement_errors = password_requirement_errors(new_password)
    if requirement_errors:
        set_flash(request, "error", " ".join(requirement_errors))
        return RedirectResponse(url="/force-password", status_code=303)

    user.password_hash = hash_password(new_password)
    user.must_change_password = False
    db.add(user)
    db.commit()
    set_flash(request, "success", "Password updated.")
    return RedirectResponse(url="/dashboard", status_code=303)


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(
    request: Request,
    location_id: list[int] | None = Query(default=None),
    provider_id: list[int] | None = Query(default=None),
    visit_date: str | None = None,
    search: str | None = None,
    hide_complete: bool = False,
    location_filter_applied: bool = False,
    provider_filter_applied: bool = False,
    db: Session = Depends(get_db),
):
    user = require_user(request, db)
    if isinstance(user, RedirectResponse):
        return user
    daily_checkout_goal = get_daily_checkout_goal(db)
    daily_checkouts_logged = sync_user_daily_checkout_state(user, db)

    locations = db.query(Location).filter(Location.is_hidden.is_(False)).order_by(Location.name).all()
    providers = db.query(Provider).filter(Provider.is_hidden.is_(False)).order_by(Provider.name).all()
    valid_location_ids = {location.id for location in locations}
    valid_provider_ids = {provider.id for provider in providers}

    selected_location_ids = normalize_selected_ids(location_id, valid_location_ids)
    selected_provider_ids = normalize_selected_ids(provider_id, valid_provider_ids)

    if not selected_location_ids and not location_filter_applied:
        selected_location_ids = persisted_selected_ids(
            user.preferred_location_ids,
            user.preferred_location_id,
            valid_location_ids,
        )
    if not selected_provider_ids and not provider_filter_applied:
        selected_provider_ids = persisted_selected_ids(
            user.preferred_provider_ids,
            user.preferred_provider_id,
            valid_provider_ids,
        )

    if location_filter_applied or provider_filter_applied or selected_location_ids or selected_provider_ids:
        persist_user_context(user, selected_location_ids, selected_provider_ids, db)

    today = date.today()
    selected_date = today
    if visit_date:
        try:
            selected_date = datetime.strptime(visit_date, "%Y-%m-%d").date()
        except ValueError:
            selected_date = today

    search_text = clamp_text(search, MAX_SEARCH_LENGTH)
    search_text_lower = search_text.lower()

    visits = []
    if selected_location_ids and selected_provider_ids:
        start_dt, end_dt = day_range(selected_date)
        visits = (
            db.query(Visit)
            .options(joinedload(Visit.location), joinedload(Visit.provider))
            .filter(
                Visit.location_id.in_(selected_location_ids),
                Visit.provider_id.in_(selected_provider_ids),
                Visit.created_at >= start_dt,
                Visit.created_at <= end_dt,
            )
            .order_by(Visit.created_at.asc())
            .all()
        )

    prepared_visits = []
    for visit in visits:
        if not visit_matches_dashboard_filters(visit, user, search_text_lower, hide_complete):
            continue

        next_field = get_next_field(visit)
        if next_field:
            action_field = next_field
            action_label = FIELD_LABELS[next_field]
            alt_action_label = None
            alt_action_field = None
            if user.role == RoleEnum.FD:
                action_enabled = next_field in {"arrived_at", "ready_for_clinical_at"}
            elif user.role == RoleEnum.NURSE:
                action_enabled = next_field in {
                    "intake_begin_at",
                    "intake_complete_at",
                    "provider_in_at",
                    "other_begin_at",
                    "other_end_at",
                    "provider_out_at",
                    "lab_complete_at",
                    "checkout_at",
                }
            else:
                action_enabled = True
        else:
            action_field = None
            action_label = "Complete"
            alt_action_label = None
            alt_action_field = None
            action_enabled = False

        prepared_visits.append(
            {
                "visit": visit,
                "status": current_status(visit),
                "next_field": action_field,
                "next_label": action_label,
                "next_enabled": action_enabled,
                "alt_action_label": alt_action_label,
                "alt_action_field": alt_action_field,
                "registration_min": minutes_between(visit.arrived_at, visit.ready_for_clinical_at),
                "waiting_room_min": minutes_between(visit.ready_for_clinical_at, visit.intake_complete_at),
                "provider_wait_min": minutes_between(visit.intake_complete_at, visit.provider_in_at),
                "provider_duration_min": minutes_between(visit.provider_in_at, visit.provider_out_at),
                "other_duration_min": other_duration_minutes(visit),
                "lab_duration_min": lab_duration_minutes(visit),
                "total_visit_min": minutes_between(visit.arrived_at, visit.checkout_at),
                "other_pending_slots": other_pending_slots(visit),
                "other_can_begin": other_can_begin(visit),
                "other_begin_options": other_begin_options(visit),
            }
        )

    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "current_user": user,
            "locations": locations,
            "providers": providers,
            "selected_location_ids": selected_location_ids,
            "selected_provider_ids": selected_provider_ids,
            "selected_location_id": selected_location_ids[0] if selected_location_ids else None,
            "selected_provider_id": selected_provider_ids[0] if selected_provider_ids else None,
            "clear_filter_query": (
                build_filter_query(selected_location_ids, selected_provider_ids, selected_date.strftime("%Y-%m-%d"), None, False)
                if selected_location_ids and selected_provider_ids
                else ""
            ),
            "selected_date": selected_date.strftime("%Y-%m-%d"),
            "search": search_text,
            "hide_complete": hide_complete,
            "visits": prepared_visits,
            "flash": pop_flash(request),
            "format_dt": format_dt,
            "selected_location_summary": summarize_selected_names(
                [location.name for location in locations if location.id in set(selected_location_ids)],
                "Choose Location",
            ),
            "selected_provider_summary": summarize_selected_names(
                [provider.name for provider in providers if provider.id in set(selected_provider_ids)],
                "Choose Provider",
            ),
            "location_picker_options": [{"value": location.id, "label": location.name} for location in locations],
            "provider_picker_options": [{"value": provider.id, "label": provider.name} for provider in providers],
            "checkout_goal_remaining": daily_checkout_goal - daily_checkouts_logged,
            "checkout_visits_logged_today": daily_checkouts_logged,
        },
    )


@app.post("/visits")
def create_visit(
    request: Request,
    mrn: str = Form(...),
    visit_type: str | None = Form(default=None),
    location_id: int = Form(...),
    provider_id: int = Form(...),
    filter_location_id: list[int] = Form(default=[]),
    filter_provider_id: list[int] = Form(default=[]),
    visit_date: str | None = Form(default=None),
    search: str | None = Form(default=None),
    hide_complete: bool = Form(default=False),
    pre_arrival: bool = Form(default=False),
    pre_arrival_date: str | None = Form(default=None),
    location_filter_applied: bool = Form(default=False),
    provider_filter_applied: bool = Form(default=False),
    db: Session = Depends(get_db),
):
    user = require_user(request, db)
    if isinstance(user, RedirectResponse):
        return user
    if user.role not in {RoleEnum.FD, RoleEnum.AUDITOR, RoleEnum.ADMIN}:
        set_flash(request, "error", "Only FD/Auditor/Admin can create visits.")
        return RedirectResponse(url="/dashboard", status_code=303)

    persisted_location_ids = filter_location_id if location_filter_applied else [location_id]
    persisted_provider_ids = filter_provider_id if provider_filter_applied else [provider_id]
    persist_user_context(user, persisted_location_ids, persisted_provider_ids, db)

    normalized_mrn = mrn.strip()
    if not normalized_mrn:
        set_flash(request, "error", "MRN is required.")
        return RedirectResponse(
            url=dashboard_redirect_url(
                persisted_location_ids,
                persisted_provider_ids,
                visit_date or date.today().strftime("%Y-%m-%d"),
                search,
                hide_complete,
                location_filter_applied,
                provider_filter_applied,
            ),
            status_code=303,
        )
    if len(normalized_mrn) > MAX_MRN_LENGTH:
        set_flash(request, "error", f"MRN must be {MAX_MRN_LENGTH} characters or fewer.")
        return RedirectResponse(
            url=dashboard_redirect_url(
                persisted_location_ids,
                persisted_provider_ids,
                visit_date or date.today().strftime("%Y-%m-%d"),
                clamp_text(search, MAX_SEARCH_LENGTH),
                hide_complete,
                location_filter_applied,
                provider_filter_applied,
            ),
            status_code=303,
        )
    if not normalized_mrn.isdigit():
        set_flash(request, "error", "MRN must contain numbers only.")
        return RedirectResponse(
            url=dashboard_redirect_url(
                persisted_location_ids,
                persisted_provider_ids,
                visit_date or date.today().strftime("%Y-%m-%d"),
                clamp_text(search, MAX_SEARCH_LENGTH),
                hide_complete,
                location_filter_applied,
                provider_filter_applied,
            ),
            status_code=303,
        )
    if not (visit_type or "").strip():
        set_flash(request, "error", "Visit Type is required.")
        return RedirectResponse(
            url=dashboard_redirect_url(
                persisted_location_ids,
                persisted_provider_ids,
                visit_date or date.today().strftime("%Y-%m-%d"),
                clamp_text(search, MAX_SEARCH_LENGTH),
                hide_complete,
                location_filter_applied,
                provider_filter_applied,
            ),
            status_code=303,
        )
    normalized_visit_type = visit_type.strip()
    if len(normalized_visit_type) > MAX_VISIT_TYPE_LENGTH:
        set_flash(request, "error", f"Visit Type must be {MAX_VISIT_TYPE_LENGTH} characters or fewer.")
        return RedirectResponse(
            url=dashboard_redirect_url(
                persisted_location_ids,
                persisted_provider_ids,
                visit_date or date.today().strftime("%Y-%m-%d"),
                clamp_text(search, MAX_SEARCH_LENGTH),
                hide_complete,
                location_filter_applied,
                provider_filter_applied,
            ),
            status_code=303,
        )

    created_at = datetime.now()
    if pre_arrival:
        scheduled_date_value = pre_arrival_date or (date.today() + timedelta(days=1)).strftime("%Y-%m-%d")
        try:
            scheduled_date = datetime.strptime(scheduled_date_value, "%Y-%m-%d").date()
            created_at = datetime.combine(scheduled_date, created_at.time())
        except ValueError:
            pass

    visit = Visit(
        mrn=normalized_mrn,
        visit_type=normalized_visit_type,
        location_id=location_id,
        provider_id=provider_id,
        arrived_at=None if pre_arrival else now_local(),
        created_by_user_id=user.id,
        created_at=created_at,
        updated_at=created_at,
    )
    db.add(visit)
    db.commit()
    set_flash(
        request,
        "success",
        f"Visit for MRN {visit.mrn} created{' as pre-arrival' if pre_arrival else ' and marked arrived'}.",
    )
    return RedirectResponse(
        url=dashboard_redirect_url(
            persisted_location_ids,
            persisted_provider_ids,
            visit_date or date.today().strftime("%Y-%m-%d"),
            clamp_text(search, MAX_SEARCH_LENGTH),
            hide_complete,
            location_filter_applied,
            provider_filter_applied,
        ),
        status_code=303,
    )




@app.get("/account", response_class=HTMLResponse)
def account_page(request: Request, db: Session = Depends(get_db)):
    user = require_user(request, db)
    if isinstance(user, RedirectResponse):
        return user
    return templates.TemplateResponse(
        "account.html",
        {
            "request": request,
            "current_user": user,
            "flash": pop_flash(request),
        },
    )


@app.post("/account/username")
def account_update_username(
    request: Request,
    username: str = Form(...),
    db: Session = Depends(get_db),
):
    user = require_user(request, db)
    if isinstance(user, RedirectResponse):
        return user

    candidate = username.strip()
    if not candidate:
        set_flash(request, "error", "Username is required.")
        return RedirectResponse(url="/account", status_code=303)
    if len(candidate) > MAX_USERNAME_LENGTH:
        set_flash(request, "error", f"Username must be {MAX_USERNAME_LENGTH} characters or fewer.")
        return RedirectResponse(url="/account", status_code=303)

    existing_user = db.query(User).filter(User.username == candidate, User.id != user.id).first()
    if existing_user:
        set_flash(request, "error", "That username is already in use.")
        return RedirectResponse(url="/account", status_code=303)

    user.username = candidate
    db.add(user)
    db.commit()
    set_flash(request, "success", "Username updated.")
    return RedirectResponse(url="/account", status_code=303)


@app.post("/account/password")
def account_update_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    db: Session = Depends(get_db),
):
    user = require_user(request, db)
    if isinstance(user, RedirectResponse):
        return user

    if not verify_password(current_password, user.password_hash):
        set_flash(request, "error", "Current password is incorrect.")
        return RedirectResponse(url="/account", status_code=303)
    if new_password != confirm_password:
        set_flash(request, "error", "New password and confirmation do not match.")
        return RedirectResponse(url="/account", status_code=303)
    requirement_errors = password_requirement_errors(new_password)
    if requirement_errors:
        set_flash(request, "error", " ".join(requirement_errors))
        return RedirectResponse(url="/account", status_code=303)

    user.password_hash = hash_password(new_password)
    user.must_change_password = False
    db.add(user)
    db.commit()
    set_flash(request, "success", "Password updated.")
    return RedirectResponse(url="/account", status_code=303)


@app.post("/account/settings")
def account_update_settings(
    request: Request,
    disable_fancy_effects: bool = Form(default=False),
    db: Session = Depends(get_db),
):
    user = require_user(request, db)
    if isinstance(user, RedirectResponse):
        return user
    user.disable_fancy_effects = disable_fancy_effects
    db.add(user)
    db.commit()
    set_flash(request, "success", "User settings updated.")
    return RedirectResponse(url="/account", status_code=303)


@app.post("/visits/{visit_id}/action")
def visit_action(
    request: Request,
    visit_id: int,
    action_field: str = Form(...),
    filter_location_id: list[int] = Form(default=[]),
    filter_provider_id: list[int] = Form(default=[]),
    visit_date: str = Form(...),
    delay_note: str | None = Form(default=None),
    other_type: str | None = Form(default=None),
    other_destination: str | None = Form(default=None),
    other_end_slot: str | None = Form(default=None),
    search: str | None = Form(default=None),
    hide_complete: bool = Form(default=False),
    location_filter_applied: bool = Form(default=False),
    provider_filter_applied: bool = Form(default=False),
    db: Session = Depends(get_db),
):
    user = require_user(request, db)
    if isinstance(user, RedirectResponse):
        return user

    persist_user_context(user, filter_location_id, filter_provider_id, db)

    visit = db.query(Visit).filter(Visit.id == visit_id).first()
    if not visit:
        set_flash(request, "error", "Visit not found.")
    else:
        try:
            set_timestamp(
                visit,
                action_field,
                user,
                db,
                delay_note=delay_note,
                other_type=other_type,
                other_destination=other_destination,
                other_end_slot=other_end_slot,
            )
            set_flash(request, "success", f"{FIELD_LABELS[action_field]} recorded for MRN {visit.mrn}.")
        except ValidationError as exc:
            set_flash(request, "error", str(exc))

    return RedirectResponse(
        url=dashboard_redirect_url(
            filter_location_id,
            filter_provider_id,
            visit_date,
            search,
            hide_complete,
            location_filter_applied,
            provider_filter_applied,
        ),
        status_code=303,
    )


@app.post("/visits/{visit_id}/declined")
def visit_declined_participation(
    request: Request,
    visit_id: int,
    declined_participation: bool = Form(default=False),
    filter_location_id: list[int] = Form(default=[]),
    filter_provider_id: list[int] = Form(default=[]),
    visit_date: str = Form(...),
    search: str | None = Form(default=None),
    hide_complete: bool = Form(default=False),
    location_filter_applied: bool = Form(default=False),
    provider_filter_applied: bool = Form(default=False),
    db: Session = Depends(get_db),
):
    user = require_user(request, db)
    if isinstance(user, RedirectResponse):
        return user

    persist_user_context(user, filter_location_id, filter_provider_id, db)

    visit = db.query(Visit).filter(Visit.id == visit_id).first()
    if not visit:
        set_flash(request, "error", "Visit not found.")
    else:
        visit.declined_participation = declined_participation
        db.add(visit)
        db.commit()

    return RedirectResponse(
        url=dashboard_redirect_url(
            filter_location_id,
            filter_provider_id,
            visit_date,
            search,
            hide_complete,
            location_filter_applied,
            provider_filter_applied,
        ),
        status_code=303,
    )


@app.post("/visits/{visit_id}/no-show")
def visit_no_show(
    request: Request,
    visit_id: int,
    no_show: bool = Form(default=False),
    filter_location_id: list[int] = Form(default=[]),
    filter_provider_id: list[int] = Form(default=[]),
    visit_date: str = Form(...),
    search: str | None = Form(default=None),
    hide_complete: bool = Form(default=False),
    location_filter_applied: bool = Form(default=False),
    provider_filter_applied: bool = Form(default=False),
    db: Session = Depends(get_db),
):
    user = require_user(request, db)
    if isinstance(user, RedirectResponse):
        return user

    persist_user_context(user, filter_location_id, filter_provider_id, db)

    visit = db.query(Visit).filter(Visit.id == visit_id).first()
    if not visit:
        set_flash(request, "error", "Visit not found.")
    else:
        visit.no_show = no_show
        db.add(visit)
        db.commit()

    return RedirectResponse(
        url=dashboard_redirect_url(
            filter_location_id,
            filter_provider_id,
            visit_date,
            search,
            hide_complete,
            location_filter_applied,
            provider_filter_applied,
        ),
        status_code=303,
    )


@app.get("/admin", response_class=HTMLResponse)
def admin_page(
    request: Request,
    mrn: str | None = None,
    search_date: str | None = None,
    field_name: str | None = None,
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    query = db.query(Visit).options(joinedload(Visit.location), joinedload(Visit.provider)).order_by(Visit.created_at.desc())
    if mrn:
        query = query.filter(Visit.mrn == mrn.strip())
    if search_date:
        try:
            d = datetime.strptime(search_date, "%Y-%m-%d").date()
            start_dt, end_dt = day_range(d)
            query = query.filter(Visit.created_at >= start_dt, Visit.created_at <= end_dt)
        except ValueError:
            pass

    visits = query.limit(100).all()
    selected_visit_id = request.query_params.get("visit_id")
    selected_field_name = field_name if field_name in TIME_FIELDS else TIME_FIELDS[0]
    audit_rows = []
    if selected_visit_id:
        audit_rows = (
            db.query(AuditLog)
            .options(joinedload(AuditLog.changed_by_user))
            .filter(AuditLog.visit_id == int(selected_visit_id))
            .order_by(AuditLog.changed_at.desc())
            .all()
        )
    admin_action_rows = (
        db.query(AdminActionLog)
        .options(joinedload(AdminActionLog.performed_by_user))
        .order_by(AdminActionLog.performed_at.desc())
        .limit(50)
        .all()
    )

    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "current_user": user,
            "visits": visits,
            "time_fields": TIME_FIELDS,
            "field_labels": FIELD_LABELS,
            "audit_rows": audit_rows,
            "selected_visit_id": int(selected_visit_id) if selected_visit_id else None,
            "selected_field_name": selected_field_name,
            "flash": pop_flash(request),
            "mrn": mrn or "",
            "search_date": search_date or "",
            "admin_action_rows": admin_action_rows,
            "format_dt": format_dt,
            "format_dt_minutes": format_dt_minutes,
            "format_dt_local_input": format_dt_local_input,
            "current_status": current_status,
            "visit_timestamp_map": {
                visit.id: {field: format_dt_local_input(getattr(visit, field)) for field in TIME_FIELDS}
                for visit in visits
            },
        },
    )


@app.post("/admin/override")
def admin_override(
    request: Request,
    visit_id: int = Form(...),
    field_name: str = Form(...),
    new_value: str = Form(...),
    reason: str = Form(...),
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    visit = db.query(Visit).filter(Visit.id == visit_id).first()
    if not visit:
        set_flash(request, "error", "Visit not found.")
        return RedirectResponse(url="/admin", status_code=303)

    parsed_new_value = None
    if new_value.strip():
        try:
            parsed_new_value = datetime.strptime(new_value, "%Y-%m-%dT%H:%M")
        except ValueError:
            set_flash(request, "error", "Invalid datetime format.")
            return RedirectResponse(url=f"/admin?visit_id={visit_id}&field_name={field_name}", status_code=303)

    try:
        override_timestamp(visit, field_name, parsed_new_value, reason, user, db)
        log_admin_action(db, user, "timestamp_override", f"Visit #{visit.id}: {field_name} overridden.")
        db.commit()
        set_flash(request, "success", "Timestamp overridden and audit log updated.")
    except ValidationError as exc:
        set_flash(request, "error", str(exc))

    return RedirectResponse(url=f"/admin?visit_id={visit_id}&field_name={field_name}", status_code=303)


@app.post("/admin/purge-zero-mrn")
def admin_purge_zero_mrn_visits(
    request: Request,
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    candidate_visits = db.query(Visit).all()
    visits_to_delete: list[Visit] = []
    for visit in candidate_visits:
        normalized_mrn = (visit.mrn or "").strip()
        if normalized_mrn and all(char == "0" for char in normalized_mrn):
            visits_to_delete.append(visit)

    deleted_visit_count = len(visits_to_delete)
    deleted_audit_log_count = 0
    if visits_to_delete:
        visit_ids = [visit.id for visit in visits_to_delete]
        deleted_audit_log_count = db.query(AuditLog).filter(AuditLog.visit_id.in_(visit_ids)).count()
        for visit in visits_to_delete:
            db.delete(visit)

    action_details = (
        f"Deleted {deleted_visit_count} visit records and {deleted_audit_log_count} audit logs "
        "for MRNs containing all zeroes."
    )
    log_admin_action(db, user, "purge_zero_mrn_records", action_details)
    db.commit()
    set_flash(request, "success", action_details)
    return RedirectResponse(url="/admin", status_code=303)


@app.get("/admin/audit-export")
def admin_audit_export(
    request: Request,
    visit_id: int | None = None,
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user
    if not visit_id:
        return Response("Visit is required for audit export.", status_code=400)

    audit_rows = (
        db.query(AuditLog)
        .options(joinedload(AuditLog.changed_by_user))
        .filter(AuditLog.visit_id == visit_id)
        .order_by(AuditLog.changed_at.desc())
        .all()
    )
    data = build_audit_export(audit_rows, FIELD_LABELS)
    log_admin_action(db, user, "visit_audit_export", f"Exported audit log for visit #{visit_id}.")
    db.commit()
    headers = {
        "Content-Disposition": f'attachment; filename="audit_visit_{visit_id}.xlsx"',
        "Cache-Control": "no-store",
    }
    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/admin/logs-export")
def admin_logs_export(
    request: Request,
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    visit_audit_rows = (
        db.query(AuditLog)
        .options(joinedload(AuditLog.changed_by_user))
        .order_by(AuditLog.changed_at.desc())
        .all()
    )
    system_audit_rows = (
        db.query(AdminActionLog)
        .options(joinedload(AdminActionLog.performed_by_user))
        .order_by(AdminActionLog.performed_at.desc())
        .all()
    )
    data = build_logs_export(visit_audit_rows, system_audit_rows, FIELD_LABELS)
    log_admin_action(
        db,
        user,
        "all_audit_logs_export",
        f"Exported {len(visit_audit_rows)} visit rows and {len(system_audit_rows)} system rows.",
    )
    db.commit()
    headers = {
        "Content-Disposition": 'attachment; filename="all_audit_logs.xlsx"',
        "Cache-Control": "no-store",
    }
    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/export", response_class=HTMLResponse)
def export_page(request: Request, db: Session = Depends(get_db)):
    user = require_user(request, db)
    if isinstance(user, RedirectResponse):
        return user

    locations = db.query(Location).filter(Location.is_hidden.is_(False)).order_by(Location.name).all()
    providers = db.query(Provider).filter(Provider.is_hidden.is_(False)).order_by(Provider.name).all()
    return templates.TemplateResponse(
        "export.html",
        {
            "request": request,
            "current_user": user,
            "locations": locations,
            "providers": providers,
            "today": date.today().strftime("%Y-%m-%d"),
        },
    )


@app.post("/export/download")
def export_download(
    request: Request,
    start_date: str = Form(...),
    end_date: str = Form(...),
    location_id: str | None = Form(default=None),
    provider_id: str | None = Form(default=None),
    include_legacy_columns: bool = Form(default=False),
    db: Session = Depends(get_db),
):
    user = require_user(request, db)
    if isinstance(user, RedirectResponse):
        return user

    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        return Response("Invalid date range.", status_code=400)
    if end < start:
        return Response("End date must be on or after start date.", status_code=400)
    if (end - start).days > MAX_EXPORT_RANGE_DAYS:
        return Response(f"Date range cannot exceed {MAX_EXPORT_RANGE_DAYS} days.", status_code=400)

    start_dt, _ = day_range(start)
    _, end_dt = day_range(end)
    try:
        parsed_location_id = int(location_id) if location_id and location_id.strip() else None
        parsed_provider_id = int(provider_id) if provider_id and provider_id.strip() else None
    except ValueError:
        return Response("Invalid location or provider filter.", status_code=400)

    query = (
        db.query(Visit)
        .options(joinedload(Visit.location), joinedload(Visit.provider), joinedload(Visit.created_by_user))
        .filter(Visit.created_at >= start_dt, Visit.created_at <= end_dt)
        .order_by(Visit.created_at.asc())
    )
    if parsed_location_id:
        query = query.filter(Visit.location_id == parsed_location_id)
    if parsed_provider_id:
        query = query.filter(Visit.provider_id == parsed_provider_id)

    try:
        data = build_export(query.all(), include_legacy_columns=include_legacy_columns)
    except Exception:
        return Response("Excel export failed.", status_code=500)

    filename = f"clinic_cycle_time_{start_date}_to_{end_date}.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Cache-Control": "no-store",
    }
    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/parameters", response_class=HTMLResponse)
def parameters_page(request: Request, db: Session = Depends(get_db)):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user
    return templates.TemplateResponse("parameters.html", parameters_page_context(request, user, db))


@app.post("/parameters/variables/goal")
def parameters_update_goal(
    request: Request,
    daily_checkout_goal: str = Form(...),
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    candidate = (daily_checkout_goal or "").strip()
    if not candidate:
        set_flash(request, "error", "Goal is required.")
        return RedirectResponse(url="/parameters", status_code=303)
    if not candidate.isdigit():
        set_flash(request, "error", "Goal must contain numbers only.")
        return RedirectResponse(url="/parameters", status_code=303)

    set_daily_checkout_goal(db, int(candidate))
    log_admin_action(db, user, "update_daily_checkout_goal", f"Set daily checkout goal to {int(candidate)}.")
    db.commit()
    set_flash(request, "success", "Daily checkout goal updated.")
    return RedirectResponse(url="/parameters", status_code=303)


@app.post("/parameters/users")
def parameters_create_user(
    request: Request,
    username: str = Form(...),
    role: RoleEnum = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    candidate = username.strip()
    if not candidate:
        set_flash(request, "error", "Username is required.")
        return RedirectResponse(url="/parameters", status_code=303)
    if len(candidate) > MAX_USERNAME_LENGTH:
        set_flash(request, "error", f"Username must be {MAX_USERNAME_LENGTH} characters or fewer.")
        return RedirectResponse(url="/parameters", status_code=303)
    if len(password) < MIN_PASSWORD_LENGTH:
        set_flash(request, "error", f"Password must be at least {MIN_PASSWORD_LENGTH} characters.")
        return RedirectResponse(url="/parameters", status_code=303)
    if len(password) > MAX_PASSWORD_LENGTH:
        set_flash(request, "error", f"Password must be {MAX_PASSWORD_LENGTH} characters or fewer.")
        return RedirectResponse(url="/parameters", status_code=303)
    if db.query(User).filter(User.username == candidate).first():
        set_flash(request, "error", "That username already exists.")
        return RedirectResponse(url="/parameters", status_code=303)

    db.add(
        User(
            username=candidate,
            password_hash=hash_password(password),
            role=role,
            is_active=True,
            must_change_password=True,
        )
    )
    log_admin_action(db, user, "create_user", f"Created user {candidate} with role {role.value}.")
    db.commit()
    set_flash(request, "success", f"User {candidate} created.")
    return RedirectResponse(url="/parameters", status_code=303)


@app.get("/parameters/users/import-guide")
def parameters_export_user_import_guide(
    request: Request,
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    data = build_user_import_guide()
    headers = {
        "Content-Disposition": 'attachment; filename="user_import_guide.xlsx"',
        "Cache-Control": "no-store",
    }
    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/parameters/users/import")
def parameters_import_users(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    filename = (file.filename or "").strip().lower()
    if not filename.endswith((".xlsx", ".xlsm")):
        set_flash(request, "error", "Please upload an .xlsx or .xlsm Excel file.")
        return RedirectResponse(url="/parameters", status_code=303)

    payload = file.file.read(MAX_USER_IMPORT_FILE_BYTES + 1)
    if len(payload) > MAX_USER_IMPORT_FILE_BYTES:
        set_flash(request, "error", "File is too large. Max size is 2MB.")
        return RedirectResponse(url="/parameters", status_code=303)

    try:
        workbook = load_workbook(BytesIO(payload), data_only=True)
        worksheet = workbook.active
    except Exception:
        set_flash(request, "error", "Invalid Excel file.")
        return RedirectResponse(url="/parameters", status_code=303)

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        set_flash(request, "error", "Excel file is empty.")
        return RedirectResponse(url="/parameters", status_code=303)

    header_index: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        normalized = str(raw or "").strip().lower()
        if normalized:
            header_index[normalized] = idx

    missing_headers = [title for title in USER_IMPORT_REQUIRED_HEADERS if title.lower() not in header_index]
    if missing_headers:
        set_flash(request, "error", f"Missing required column(s): {', '.join(missing_headers)}.")
        return RedirectResponse(url="/parameters", status_code=303)

    username_idx = header_index["username"]
    role_idx = header_index["role"]
    password_idx = header_index["temporary password"]

    existing_usernames = {name for (name,) in db.query(User.username).all()}
    seen_in_file: set[str] = set()
    created_usernames: list[str] = []
    skipped_existing = 0
    row_errors: list[str] = []

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if row_number > MAX_USER_IMPORT_ROWS + 1:
            row_errors.append(f"Row {row_number}: too many rows (max {MAX_USER_IMPORT_ROWS}).")
            break
        if not any(value is not None and str(value).strip() for value in row):
            continue

        username = str(row[username_idx] or "").strip()
        role_value = str(row[role_idx] or "").strip().lower()
        password = str(row[password_idx] or "")

        if not username:
            row_errors.append(f"Row {row_number}: Username is required.")
            continue
        if len(username) > MAX_USERNAME_LENGTH:
            row_errors.append(f"Row {row_number}: Username exceeds {MAX_USERNAME_LENGTH} characters.")
            continue
        if len(password) < MIN_PASSWORD_LENGTH:
            row_errors.append(f"Row {row_number}: Temporary Password must be at least {MIN_PASSWORD_LENGTH} characters.")
            continue
        if len(password) > MAX_PASSWORD_LENGTH:
            row_errors.append(f"Row {row_number}: Temporary Password exceeds {MAX_PASSWORD_LENGTH} characters.")
            continue

        try:
            parsed_role = RoleEnum(role_value)
        except Exception:
            row_errors.append(f"Row {row_number}: Role must be one of admin, fd, nurse, auditor.")
            continue

        if username in existing_usernames or username in seen_in_file:
            skipped_existing += 1
            continue

        db.add(
            User(
                username=username,
                password_hash=hash_password(password),
                role=parsed_role,
                is_active=True,
                must_change_password=True,
            )
        )
        seen_in_file.add(username)
        created_usernames.append(username)

    if created_usernames:
        sample_users = ", ".join(created_usernames[:10])
        detail_suffix = "" if len(created_usernames) <= 10 else f" (+{len(created_usernames) - 10} more)"
        log_admin_action(
            db,
            user,
            "import_users",
            f"Imported {len(created_usernames)} users: {sample_users}{detail_suffix}. "
            f"Skipped existing/duplicate: {skipped_existing}. Errors: {len(row_errors)}.",
        )
        db.commit()

    if row_errors:
        error_preview = " ".join(row_errors[:3])
        if len(row_errors) > 3:
            error_preview += f" (+{len(row_errors) - 3} more errors)"
        if created_usernames:
            set_flash(
                request,
                "success",
                f"Imported {len(created_usernames)} users. Skipped {skipped_existing}. {error_preview}",
            )
        else:
            set_flash(request, "error", f"No users imported. {error_preview}")
        return RedirectResponse(url="/parameters", status_code=303)

    if created_usernames:
        set_flash(
            request,
            "success",
            f"Imported {len(created_usernames)} users. Skipped {skipped_existing} existing/duplicate users.",
        )
    else:
        set_flash(request, "error", "No users imported. All rows were duplicates or empty.")
    return RedirectResponse(url="/parameters", status_code=303)


@app.post("/parameters/users/{user_id}/reset-password")
def parameters_reset_password(
    request: Request,
    user_id: int,
    new_password: str = Form(...),
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    target_user = db.query(User).filter(User.id == user_id).first()
    if not target_user:
        set_flash(request, "error", "User not found.")
        return RedirectResponse(url="/parameters", status_code=303)
    if len(new_password) < MIN_PASSWORD_LENGTH:
        set_flash(request, "error", f"Password must be at least {MIN_PASSWORD_LENGTH} characters.")
        return RedirectResponse(url="/parameters", status_code=303)
    if len(new_password) > MAX_PASSWORD_LENGTH:
        set_flash(request, "error", f"Password must be {MAX_PASSWORD_LENGTH} characters or fewer.")
        return RedirectResponse(url="/parameters", status_code=303)

    target_user.password_hash = hash_password(new_password)
    target_user.must_change_password = True
    db.add(target_user)
    log_admin_action(db, user, "reset_user_password", f"Reset password for {target_user.username}.")
    db.commit()
    set_flash(request, "success", f"Password reset for {target_user.username}.")
    return RedirectResponse(url="/parameters", status_code=303)


@app.post("/parameters/users/{user_id}/role")
def parameters_update_user_role(
    request: Request,
    user_id: int,
    role: RoleEnum = Form(...),
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    target_user = db.query(User).filter(User.id == user_id).first()
    if not target_user:
        set_flash(request, "error", "User not found.")
        return RedirectResponse(url="/parameters", status_code=303)

    target_user.role = role
    db.add(target_user)
    log_admin_action(db, user, "update_user_role", f"Updated role for {target_user.username} to {role.value}.")
    db.commit()
    set_flash(request, "success", f"Role updated for {target_user.username}.")
    return RedirectResponse(url="/parameters", status_code=303)


@app.post("/parameters/users/{user_id}/hidden")
def parameters_update_user_hidden(
    request: Request,
    user_id: int,
    is_hidden: bool = Form(default=False),
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    target_user = db.query(User).filter(User.id == user_id).first()
    if not target_user:
        set_flash(request, "error", "User not found.")
        return RedirectResponse(url="/parameters", status_code=303)

    target_user.is_hidden = is_hidden
    db.add(target_user)
    log_admin_action(db, user, "update_user_hidden", f"Set hidden={is_hidden} for {target_user.username}.")
    db.commit()
    set_flash(request, "success", f"Hidden updated for {target_user.username}.")
    return RedirectResponse(url="/parameters", status_code=303)


@app.post("/parameters/locations")
def parameters_add_location(
    request: Request,
    name: str = Form(...),
    is_hidden: bool = Form(default=False),
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    candidate = name.strip()
    if not candidate:
        set_flash(request, "error", "Location name is required.")
        return RedirectResponse(url="/parameters", status_code=303)
    if len(candidate) > MAX_LOCATION_NAME_LENGTH:
        set_flash(request, "error", f"Location name must be {MAX_LOCATION_NAME_LENGTH} characters or fewer.")
        return RedirectResponse(url="/parameters", status_code=303)
    if db.query(Location).filter(Location.name == candidate).first():
        set_flash(request, "error", "That location already exists.")
        return RedirectResponse(url="/parameters", status_code=303)

    db.add(Location(name=candidate, is_hidden=is_hidden))
    log_admin_action(db, user, "add_location", f"Added location {candidate} (hidden={is_hidden}).")
    db.commit()
    set_flash(request, "success", f"Location {candidate} added.")
    return RedirectResponse(url="/parameters", status_code=303)


@app.post("/parameters/locations/{location_id}")
def parameters_update_location(
    request: Request,
    location_id: int,
    name: str = Form(...),
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    location = db.query(Location).filter(Location.id == location_id).first()
    candidate = name.strip()
    if not location:
        set_flash(request, "error", "Location not found.")
        return RedirectResponse(url="/parameters", status_code=303)
    if not candidate:
        set_flash(request, "error", "Location name is required.")
        return RedirectResponse(url="/parameters", status_code=303)
    if len(candidate) > MAX_LOCATION_NAME_LENGTH:
        set_flash(request, "error", f"Location name must be {MAX_LOCATION_NAME_LENGTH} characters or fewer.")
        return RedirectResponse(url="/parameters", status_code=303)
    if db.query(Location).filter(Location.name == candidate, Location.id != location_id).first():
        set_flash(request, "error", "That location name is already in use.")
        return RedirectResponse(url="/parameters", status_code=303)

    location.name = candidate
    db.add(location)
    log_admin_action(db, user, "update_location", f"Updated location #{location.id} to {candidate}.")
    db.commit()
    set_flash(request, "success", "Location updated.")
    return RedirectResponse(url="/parameters", status_code=303)


@app.post("/parameters/locations/{location_id}/hidden")
def parameters_update_location_hidden(
    request: Request,
    location_id: int,
    is_hidden: bool = Form(default=False),
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    location = db.query(Location).filter(Location.id == location_id).first()
    if not location:
        set_flash(request, "error", "Location not found.")
        return RedirectResponse(url="/parameters", status_code=303)

    location.is_hidden = is_hidden
    db.add(location)
    log_admin_action(db, user, "update_location_hidden", f"Set hidden={is_hidden} for location #{location.id}.")
    db.commit()
    set_flash(request, "success", "Location hidden state updated.")
    return RedirectResponse(url="/parameters", status_code=303)


@app.post("/parameters/providers")
def parameters_add_provider(
    request: Request,
    name: str = Form(...),
    is_hidden: bool = Form(default=False),
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    candidate = name.strip()
    if not candidate:
        set_flash(request, "error", "Provider name is required.")
        return RedirectResponse(url="/parameters", status_code=303)
    if len(candidate) > MAX_PROVIDER_NAME_LENGTH:
        set_flash(request, "error", f"Provider name must be {MAX_PROVIDER_NAME_LENGTH} characters or fewer.")
        return RedirectResponse(url="/parameters", status_code=303)
    if db.query(Provider).filter(Provider.name == candidate).first():
        set_flash(request, "error", "That provider already exists.")
        return RedirectResponse(url="/parameters", status_code=303)

    db.add(Provider(name=candidate, is_hidden=is_hidden))
    log_admin_action(db, user, "add_provider", f"Added provider {candidate} (hidden={is_hidden}).")
    db.commit()
    set_flash(request, "success", f"Provider {candidate} added.")
    return RedirectResponse(url="/parameters", status_code=303)


@app.post("/parameters/providers/{provider_id}")
def parameters_update_provider(
    request: Request,
    provider_id: int,
    name: str = Form(...),
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    provider = db.query(Provider).filter(Provider.id == provider_id).first()
    candidate = name.strip()
    if not provider:
        set_flash(request, "error", "Provider not found.")
        return RedirectResponse(url="/parameters", status_code=303)
    if not candidate:
        set_flash(request, "error", "Provider name is required.")
        return RedirectResponse(url="/parameters", status_code=303)
    if len(candidate) > MAX_PROVIDER_NAME_LENGTH:
        set_flash(request, "error", f"Provider name must be {MAX_PROVIDER_NAME_LENGTH} characters or fewer.")
        return RedirectResponse(url="/parameters", status_code=303)
    if db.query(Provider).filter(Provider.name == candidate, Provider.id != provider_id).first():
        set_flash(request, "error", "That provider name is already in use.")
        return RedirectResponse(url="/parameters", status_code=303)

    provider.name = candidate
    db.add(provider)
    log_admin_action(db, user, "update_provider", f"Updated provider #{provider.id} to {candidate}.")
    db.commit()
    set_flash(request, "success", "Provider updated.")
    return RedirectResponse(url="/parameters", status_code=303)


@app.post("/parameters/providers/{provider_id}/hidden")
def parameters_update_provider_hidden(
    request: Request,
    provider_id: int,
    is_hidden: bool = Form(default=False),
    db: Session = Depends(get_db),
):
    user = require_admin(request, db)
    if isinstance(user, RedirectResponse):
        return user

    provider = db.query(Provider).filter(Provider.id == provider_id).first()
    if not provider:
        set_flash(request, "error", "Provider not found.")
        return RedirectResponse(url="/parameters", status_code=303)

    provider.is_hidden = is_hidden
    db.add(provider)
    log_admin_action(db, user, "update_provider_hidden", f"Set hidden={is_hidden} for provider #{provider.id}.")
    db.commit()
    set_flash(request, "success", "Provider hidden state updated.")
    return RedirectResponse(url="/parameters", status_code=303)


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("app.main:app", host="0.0.0.0", port=8000)
