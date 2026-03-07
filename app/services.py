from __future__ import annotations

import json
from datetime import date, datetime, time
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from sqlalchemy import update
from sqlalchemy.orm import Session

from app.models import AuditLog, RoleEnum, User, Visit

TIME_FIELDS = [
    "arrived_at",
    "ready_for_clinical_at",
    "intake_begin_at",
    "intake_complete_at",
    "provider_in_at",
    "other_begin_at",
    "other_end_at",
    "provider_out_at",
    "lab_complete_at",
    "checkout_at",
]

FIELD_LABELS = {
    "arrived_at": "Arrived",
    "ready_for_clinical_at": "Ready for Clinical",
    "intake_begin_at": "Intake Begin",
    "intake_complete_at": "Intake Complete",
    "provider_in_at": "Provider In",
    "other_begin_at": "Other Begin",
    "other_end_at": "Other End",
    "provider_out_at": "Provider Out",
    "lab_complete_at": "Lab Complete",
    "checkout_at": "Checkout",
}

DELAY_NOTE_FIELDS = {
    "arrived_at": "arrived_delay_note",
    "ready_for_clinical_at": "ready_for_clinical_delay_note",
    "intake_begin_at": "intake_begin_delay_note",
    "intake_complete_at": "intake_complete_delay_note",
    "provider_in_at": "provider_in_delay_note",
    "other_begin_at": "other_begin_delay_note",
    "other_end_at": "other_end_delay_note",
    "provider_out_at": "provider_out_delay_note",
    "lab_complete_at": "lab_complete_delay_note",
    "checkout_at": "checkout_delay_note",
}

ROLE_ACTIONS = {
    RoleEnum.FD: {"arrived_at", "ready_for_clinical_at"},
    RoleEnum.NURSE: {
        "intake_begin_at",
        "intake_complete_at",
        "provider_in_at",
        "other_begin_at",
        "other_end_at",
        "provider_out_at",
        "lab_complete_at",
        "checkout_at",
    },
    RoleEnum.AUDITOR: set(TIME_FIELDS),
    RoleEnum.ADMIN: set(TIME_FIELDS),
}

OTHER_SLOT_ORDER = ["Lab", "Ultrasound", "X-Ray", "Other 1", "Other 2"]


class ValidationError(Exception):
    pass


def now_local() -> datetime:
    return datetime.now()


def can_set_field(visit: Visit, field_name: str) -> bool:
    if getattr(visit, field_name) is not None:
        return False

    if field_name == "arrived_at":
        return True
    if field_name == "ready_for_clinical_at":
        return visit.arrived_at is not None
    if field_name == "intake_begin_at":
        return visit.ready_for_clinical_at is not None
    if field_name == "intake_complete_at":
        return visit.intake_begin_at is not None
    if field_name == "provider_in_at":
        return visit.intake_complete_at is not None
    if field_name == "other_begin_at":
        return visit.intake_complete_at is not None
    if field_name == "other_end_at":
        return visit.other_begin_at is not None
    if field_name == "provider_out_at":
        return visit.provider_in_at is not None
    if field_name == "lab_complete_at":
        return visit.intake_complete_at is not None
    if field_name == "checkout_at":
        if visit.provider_out_at is None:
            return False
        return True
    return False


def get_next_field(visit: Visit) -> str | None:
    if visit.arrived_at is None:
        return "arrived_at"
    if visit.ready_for_clinical_at is None:
        return "ready_for_clinical_at"
    if visit.intake_begin_at is None:
        return "intake_begin_at"
    if visit.intake_complete_at is None:
        return "intake_complete_at"
    if visit.provider_in_at is None:
        return "provider_in_at"
    if visit.provider_out_at is None:
        return "provider_out_at"
    if visit.checkout_at is None:
        return "checkout_at"
    return None


def current_status(visit: Visit) -> str:
    next_field = get_next_field(visit)
    if next_field is None:
        return "Complete"
    if next_field == "arrived_at":
        return "Awaiting Arrival"
    return f"Awaiting {FIELD_LABELS[next_field]}"


def lab_duration_minutes(visit: Visit) -> float | None:
    if not visit.lab_complete_at:
        return None
    if visit.provider_out_at and visit.provider_out_at <= visit.lab_complete_at:
        return minutes_between(visit.provider_out_at, visit.lab_complete_at)
    if visit.provider_in_at and visit.provider_in_at <= visit.lab_complete_at:
        return minutes_between(visit.provider_in_at, visit.lab_complete_at)
    if visit.intake_complete_at:
        return minutes_between(visit.intake_complete_at, visit.lab_complete_at)
    return None


def other_duration_minutes(visit: Visit) -> float | None:
    return minutes_between(visit.other_begin_at, visit.other_end_at)


def _parse_other_tracking(visit: Visit) -> list[dict]:
    raw_value = (visit.other_timestamps_json or "").strip()
    if not raw_value:
        return []
    try:
        parsed = json.loads(raw_value)
        return parsed if isinstance(parsed, list) else []
    except json.JSONDecodeError:
        return []


def _serialize_other_tracking(entries: list[dict]) -> str:
    return json.dumps(entries, separators=(",", ":"))


def _other_slot_values(visit: Visit) -> dict[str, dict]:
    values: dict[str, dict] = {}
    for entry in _parse_other_tracking(visit):
        slot = str(entry.get("slot", "")).strip()
        if slot:
            values[slot] = entry
    return values


def _other_logs(entries: list[dict]) -> tuple[str | None, str | None]:
    begin_lines: list[str] = []
    end_lines: list[str] = []
    for entry in entries:
        slot = entry.get("slot")
        begin_at = entry.get("begin_at")
        end_at = entry.get("end_at")
        begin_note = entry.get("begin_note")
        end_note = entry.get("end_note")
        if slot and begin_at:
            line = f"{begin_at} - {slot}"
            if begin_note:
                line = f"{line} ({begin_note})"
            begin_lines.append(line)
        if slot and end_at:
            line = f"{end_at} - {slot}"
            if end_note:
                line = f"{line} ({end_note})"
            end_lines.append(line)
    return ("\n".join(begin_lines) if begin_lines else None, "\n".join(end_lines) if end_lines else None)


def other_pending_slots(visit: Visit) -> list[dict[str, str]]:
    pending: list[dict[str, str]] = []
    for entry in _parse_other_tracking(visit):
        slot = str(entry.get("slot", "")).strip()
        begin_at = str(entry.get("begin_at", "")).strip()
        end_at = str(entry.get("end_at", "")).strip()
        begin_note = str(entry.get("begin_note", "")).strip()
        if slot and begin_at and not end_at:
            pending.append({"value": slot, "label": begin_note or slot})
    return pending


def other_begin_options(visit: Visit) -> list[dict[str, str]]:
    entries = _parse_other_tracking(visit)
    used_slots = {str(entry.get("slot", "")).strip() for entry in entries}
    options: list[dict[str, str]] = []
    for value in ["Lab", "X-Ray", "Ultrasound"]:
        if value not in used_slots:
            options.append({"value": value, "label": value})
    if "Other 1" not in used_slots or "Other 2" not in used_slots:
        options.append({"value": "OTHER", "label": "OTHER"})
    return options


def other_can_begin(visit: Visit) -> bool:
    return len(_parse_other_tracking(visit)) < 5 and bool(other_begin_options(visit))


def delay_note_entries(visit: Visit) -> list[tuple[str, str]]:
    entries: list[tuple[str, str]] = []
    for field_name in TIME_FIELDS:
        note_value = getattr(visit, DELAY_NOTE_FIELDS[field_name])
        if note_value:
            entries.append((FIELD_LABELS[field_name], note_value))
    return entries


def set_timestamp(
    visit: Visit,
    field_name: str,
    acting_user: User,
    db: Session,
    delay_note: str | None = None,
    other_type: str | None = None,
    other_destination: str | None = None,
    other_end_slot: str | None = None,
) -> None:
    if field_name not in TIME_FIELDS:
        raise ValidationError("Invalid field.")

    if field_name not in ROLE_ACTIONS[acting_user.role]:
        raise ValidationError("You are not allowed to perform this action.")

    now_value = now_local()
    timestamp_label = now_value.strftime("%Y-%m-%d %H:%M:%S")
    existing_note = getattr(visit, DELAY_NOTE_FIELDS[field_name])
    new_note = delay_note.strip() if delay_note and delay_note.strip() else None
    tracking_entries = _parse_other_tracking(visit)
    tracking_changed = False
    if field_name == "other_begin_at":
        selected_type = (other_type or "").strip()
        allowed_types = {"Lab", "X-Ray", "Ultrasound", "OTHER"}
        if selected_type not in allowed_types:
            raise ValidationError("Choose Lab, X-Ray, Ultrasound, or OTHER for Other Begin.")
        typed_destination = (other_destination or "").strip()
        if selected_type == "OTHER" and not typed_destination:
            raise ValidationError("When OTHER is selected, enter where the patient was sent.")
        if len(tracking_entries) >= 5:
            raise ValidationError("Maximum of 5 Other Begin timestamps reached for this visit.")

        used_slots = {str(entry.get("slot", "")).strip() for entry in tracking_entries}
        if selected_type in {"Lab", "X-Ray", "Ultrasound"}:
            slot = selected_type
            if slot in used_slots:
                raise ValidationError(f"{slot} already started for this visit.")
        else:
            slot = "Other 1" if "Other 1" not in used_slots else ("Other 2" if "Other 2" not in used_slots else "")
            if not slot:
                raise ValidationError("Maximum OTHER entries reached for this visit.")

        begin_description = selected_type if selected_type != "OTHER" else typed_destination
        tracking_entries.append(
            {
                "slot": slot,
                "type": selected_type,
                "begin_at": timestamp_label,
                "begin_note": begin_description,
                "end_at": "",
                "end_note": "",
            }
        )
        tracking_changed = True
        note_value = begin_description
    elif field_name == "other_end_at":
        selected_slot = (other_end_slot or "").strip()
        if not selected_slot:
            raise ValidationError("Select a destination to close for Other End.")
        matching_entry = None
        for entry in tracking_entries:
            if str(entry.get("slot", "")).strip() == selected_slot and not str(entry.get("end_at", "")).strip():
                matching_entry = entry
                break
        if not matching_entry:
            raise ValidationError("Selected Other destination is not currently open.")
        matching_entry["end_at"] = timestamp_label
        matching_entry["end_note"] = new_note or ""
        tracking_changed = True
        note_value = f"{existing_note}; {new_note}" if existing_note and new_note else (new_note or existing_note)
    else:
        note_value = new_note

    other_begin_log_value = visit.other_begin_log
    other_end_log_value = visit.other_end_log
    other_json_value = visit.other_timestamps_json
    if tracking_changed:
        other_json_value = _serialize_other_tracking(tracking_entries)
        other_begin_log_value, other_end_log_value = _other_logs(tracking_entries)

    update_values = {
        field_name: now_value,
        DELAY_NOTE_FIELDS[field_name]: note_value,
        "updated_at": now_value,
    }
    if field_name == "arrived_at":
        # Track ownership by the user who records the first Arrived timestamp.
        update_values["created_by_user_id"] = acting_user.id
    if field_name == "other_begin_at":
        update_values["other_begin_log"] = other_begin_log_value
    elif field_name == "other_end_at":
        update_values["other_end_log"] = other_end_log_value
    if tracking_changed:
        update_values["other_timestamps_json"] = other_json_value
    conditions = [Visit.id == visit.id]
    if field_name not in {"other_begin_at", "other_end_at"}:
        conditions.append(getattr(Visit, field_name).is_(None))

    if field_name == "ready_for_clinical_at":
        conditions.append(Visit.arrived_at.is_not(None))
    elif field_name == "intake_begin_at":
        conditions.append(Visit.ready_for_clinical_at.is_not(None))
    elif field_name == "intake_complete_at":
        conditions.append(Visit.intake_begin_at.is_not(None))
    elif field_name == "provider_in_at":
        conditions.append(Visit.intake_complete_at.is_not(None))
    elif field_name == "other_begin_at":
        conditions.append(Visit.intake_complete_at.is_not(None))
    elif field_name == "other_end_at":
        conditions.append(Visit.other_begin_at.is_not(None))
    elif field_name == "provider_out_at":
        conditions.append(Visit.provider_in_at.is_not(None))
    elif field_name == "lab_complete_at":
        conditions.append(Visit.intake_complete_at.is_not(None))
    elif field_name == "checkout_at":
        conditions.append(Visit.provider_out_at.is_not(None))

    result = db.execute(update(Visit).where(*conditions).values(**update_values))
    if result.rowcount != 1:
        db.rollback()
        current_visit = db.query(Visit).filter(Visit.id == visit.id).first()
        if not current_visit:
            raise ValidationError("Visit not found.")
        db.refresh(current_visit)
        for attribute in TIME_FIELDS:
            setattr(visit, attribute, getattr(current_visit, attribute))
        for attribute in DELAY_NOTE_FIELDS.values():
            setattr(visit, attribute, getattr(current_visit, attribute))
        visit.other_timestamps_json = current_visit.other_timestamps_json
        visit.other_begin_log = current_visit.other_begin_log
        visit.other_end_log = current_visit.other_end_log
        visit.updated_at = current_visit.updated_at
        if field_name not in {"other_begin_at", "other_end_at"} and getattr(current_visit, field_name) is not None:
            raise ValidationError("This step was already recorded by another user. The queue has been refreshed.")
        raise ValidationError("This visit changed before your save completed. The queue has been refreshed.")

    db.commit()
    db.refresh(visit)


def override_timestamp(
    visit: Visit,
    field_name: str,
    new_value: datetime | None,
    reason: str,
    acting_user: User,
    db: Session,
) -> None:
    if acting_user.role != RoleEnum.ADMIN:
        raise ValidationError("Only admins can override timestamps.")
    if field_name not in TIME_FIELDS:
        raise ValidationError("Invalid timestamp field.")
    if not reason.strip():
        raise ValidationError("Reason is required.")

    old_value = getattr(visit, field_name)
    setattr(visit, field_name, new_value)
    visit.updated_at = now_local()

    audit = AuditLog(
        visit_id=visit.id,
        field_name=field_name,
        old_value=format_dt(old_value),
        new_value=format_dt(new_value),
        changed_by_user_id=acting_user.id,
        changed_at=now_local(),
        reason=reason.strip(),
    )
    db.add(visit)
    db.add(audit)
    db.commit()


def format_dt(value: datetime | None) -> str | None:
    if value is None:
        return None
    return value.strftime("%Y-%m-%d %H:%M:%S")


def minutes_between(start: datetime | None, end: datetime | None) -> float | None:
    if not start or not end:
        return None
    return round((end - start).total_seconds() / 60, 2)


def build_export(
    visits: list[Visit],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "CycleTime"

    headers = [
        "MRN",
        "Location",
        "Provider",
        "Visit Type",
        "Arrived At",
        "Arrived Recorded By",
        "Arrived Delay Reason",
        "Ready for Clinical At",
        "Ready for Clinical Delay Reason",
        "Intake Begin At",
        "Intake Begin Delay Reason",
        "Intake Complete At",
        "Intake Complete Delay Reason",
        "Provider In At",
        "Provider In Delay Reason",
        "Lab Begin At",
        "Lab Begin Note",
        "Lab End At",
        "Lab End Note",
        "Ultrasound Begin At",
        "Ultrasound Begin Note",
        "Ultrasound End At",
        "Ultrasound End Note",
        "X-Ray Begin At",
        "X-Ray Begin Note",
        "X-Ray End At",
        "X-Ray End Note",
        "Other 1 Begin At",
        "Other 1 Location",
        "Other 1 End At",
        "Other 1 End Note",
        "Other 2 Begin At",
        "Other 2 Location",
        "Other 2 End At",
        "Other 2 End Note",
        "Other Begin Delay Reason (Legacy)",
        "Other End Delay Reason (Legacy)",
        "Provider Out At",
        "Provider Out Delay Reason",
        "Lab Complete At",
        "Lab Complete Delay Reason",
        "Checkout At",
        "Checkout Delay Reason",
        "Registration Minutes",
        "Waiting Room Minutes",
        "Provider Wait Minutes",
        "Provider Duration Minutes",
        "Other Duration Minutes",
        "Lab Duration Minutes",
        "Total Visit Minutes",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for column_index in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(column_index)].width = 18.0

    for visit in visits:
        slot_values = _other_slot_values(visit)
        ws.append(
            [
                visit.mrn,
                visit.location.name,
                visit.provider.name,
                visit.visit_type,
                format_dt(visit.arrived_at),
                visit.created_by_user.username if visit.created_by_user else "",
                visit.arrived_delay_note,
                format_dt(visit.ready_for_clinical_at),
                visit.ready_for_clinical_delay_note,
                format_dt(visit.intake_begin_at),
                visit.intake_begin_delay_note,
                format_dt(visit.intake_complete_at),
                visit.intake_complete_delay_note,
                format_dt(visit.provider_in_at),
                visit.provider_in_delay_note,
                slot_values.get("Lab", {}).get("begin_at"),
                slot_values.get("Lab", {}).get("begin_note"),
                slot_values.get("Lab", {}).get("end_at"),
                slot_values.get("Lab", {}).get("end_note"),
                slot_values.get("Ultrasound", {}).get("begin_at"),
                slot_values.get("Ultrasound", {}).get("begin_note"),
                slot_values.get("Ultrasound", {}).get("end_at"),
                slot_values.get("Ultrasound", {}).get("end_note"),
                slot_values.get("X-Ray", {}).get("begin_at"),
                slot_values.get("X-Ray", {}).get("begin_note"),
                slot_values.get("X-Ray", {}).get("end_at"),
                slot_values.get("X-Ray", {}).get("end_note"),
                slot_values.get("Other 1", {}).get("begin_at"),
                slot_values.get("Other 1", {}).get("begin_note"),
                slot_values.get("Other 1", {}).get("end_at"),
                slot_values.get("Other 1", {}).get("end_note"),
                slot_values.get("Other 2", {}).get("begin_at"),
                slot_values.get("Other 2", {}).get("begin_note"),
                slot_values.get("Other 2", {}).get("end_at"),
                slot_values.get("Other 2", {}).get("end_note"),
                visit.other_begin_delay_note,
                visit.other_end_delay_note,
                format_dt(visit.provider_out_at),
                visit.provider_out_delay_note,
                format_dt(visit.lab_complete_at),
                visit.lab_complete_delay_note,
                format_dt(visit.checkout_at),
                visit.checkout_delay_note,
                minutes_between(visit.arrived_at, visit.ready_for_clinical_at),
                minutes_between(visit.ready_for_clinical_at, visit.intake_complete_at),
                minutes_between(visit.intake_complete_at, visit.provider_in_at),
                minutes_between(visit.provider_in_at, visit.provider_out_at),
                other_duration_minutes(visit),
                lab_duration_minutes(visit),
                minutes_between(visit.arrived_at, visit.checkout_at),
            ]
        )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def build_audit_export(audit_rows: list[AuditLog], field_labels: dict[str, str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "AuditLog"

    headers = ["Changed At", "User", "Field", "Old", "New", "Reason"]
    ws.append(headers)
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for column_index in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(column_index)].width = 24.0

    for row in audit_rows:
        ws.append(
            [
                format_dt(row.changed_at),
                row.changed_by_user.username if row.changed_by_user else "",
                field_labels.get(row.field_name, row.field_name),
                row.old_value or "",
                row.new_value or "",
                row.reason,
            ]
        )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def day_range(day: date) -> tuple[datetime, datetime]:
    return datetime.combine(day, time.min), datetime.combine(day, time.max)
